from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.user import User
from app.models.project import Project
from app.models.post import Post
from app.models.brand import Brand

router = APIRouter()

@router.get("/excel/{project_id}")
async def export_project_excel(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Esporta il calendario in Excel"""
    
    # Verifica accesso
    project = db.query(Project).join(Brand).filter(
        Project.id == project_id,
        Brand.organization_id == current_user.organization_id
    ).first()
    
    if not project:
        raise HTTPException(status_code=404, detail="Progetto non trovato")
    
    brand = db.query(Brand).filter(Brand.id == project.brand_id).first()
    posts = db.query(Post).filter(Post.project_id == project_id).order_by(Post.scheduled_date, Post.scheduled_time).all()
    
    # Crea workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Piano Editoriale"
    
    # Stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3DAFA8", end_color="3DAFA8", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header info progetto
    ws['A1'] = f"Piano Editoriale: {project.name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Brand: {brand.name if brand else 'N/A'}"
    ws['A3'] = f"Periodo: {project.start_date} - {project.end_date}"
    ws['A4'] = f"Esportato il: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    
    # Headers tabella
    headers = ["Data", "Ora", "Piattaforma", "Contenuto", "Hashtags", "Pillar", "Tipo", "CTA", "Visual", "Status"]
    row = 6
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Colori piattaforme
    platform_colors = {
        'linkedin': 'E3F2FD',
        'instagram': 'FCE4EC',
        'facebook': 'E8EAF6',
        'google_business': 'E8F5E9'
    }
    
    # Dati
    for post in posts:
        row += 1
        data = [
            post.scheduled_date.strftime('%d/%m/%Y') if post.scheduled_date else '',
            post.scheduled_time or '',
            post.platform.upper() if post.platform else '',
            post.content or '',
            ', '.join(post.hashtags) if post.hashtags else '',
            post.pillar or '',
            post.post_type or '',
            post.cta or '',
            post.visual_suggestion or '',
            post.status or 'draft'
        ]
        
        fill_color = platform_colors.get(post.platform, 'FFFFFF')
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.fill = row_fill
            if col == 4:  # Contenuto
                cell.alignment = Alignment(wrap_text=True)
    
    # Larghezza colonne
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 10
    
    # Salva in memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"piano_editoriale_{project.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
